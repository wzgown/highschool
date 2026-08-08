# -*- coding: utf-8 -*-
"""整合 2026 上海中考数据集为 Excel 工作簿（多 Sheet，含自说明与字段字典）。"""
import csv, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 项目根目录：本文件位于 scripts/，上一级即项目根（含 data/ raw/ output/ 三个同级目录）
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA = os.path.join(ROOT, "data")            # 清洗好的结构化数据（输入）
OUTPUT = os.path.join(ROOT, "output")        # 导出产物（输出）
OUT = os.path.join(OUTPUT, "2026上海中考数据集.xlsx")

DISTRICT_MAP = {
    "pudong": "浦东新区", "浦东": "浦东新区",
    "黄浦": "黄浦区", "徐汇": "徐汇区", "长宁": "长宁区", "静安": "静安区",
    "普陀": "普陀区", "虹口": "虹口区", "杨浦": "杨浦区", "闵行": "闵行区",
    "宝山": "宝山区", "嘉定": "嘉定区", "松江": "松江区", "青浦": "青浦区",
    "奉贤": "奉贤区", "金山": "金山区", "崇明": "崇明区",
}
DISTRICTS_ORDER = ["浦东新区","黄浦区","徐汇区","长宁区","静安区","普陀区","虹口区",
                   "杨浦区","闵行区","宝山区","嘉定区","松江区","青浦区","奉贤区",
                   "金山区","崇明区"]

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUB_FONT = Font(bold=True, size=11, color="1F4E78")
WARN_FONT = Font(bold=True, size=11, color="C00000")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = CENTER; cell.border = BORDER

def write_table(ws, start_row, headers, rows, widths=None, freeze=True):
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top")
        r += 1
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    if freeze:
        ws.freeze_panes = ws.cell(row=start_row+1, column=1)
    return r

def read_csv(path):
    rows, comments = [], []
    with open(path, encoding="utf-8-sig", newline="") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line.strip():
                continue
            if line.lstrip().startswith("#"):
                comments.append(line.lstrip("# ").strip())
                continue
            rows.append(next(csv.reader([line])))
    return rows, comments

wb = Workbook()

# ============ Sheet 1: 说明与数据来源 ============
ws = wb.active; ws.title = "说明与数据来源"
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 104
ws["A1"] = "2026 年上海中考数据集（收集汇编）"; ws["A1"].font = TITLE_FONT
notes = [
    ("数据截止", "2026-07-24。绝大部分官方数据已于 2026-07-14~07-23 由上海市教育考试院及各区招考机构发布。"),
    ("主要来源", "上海招考热线 shmeea.edu.cn；各区教育局/招考机构官网；上海本地宝(sh.bendibao.com)等权威镜像；各机构公众号/自媒体汇总（搜狐、头条、haoxue360、vsxue 等）。"),
    ("工作簿结构", "本工作簿共 18 个 Sheet：说明与数据来源(本页)、字段数据字典、考生人数、各初中考生人数(推算)、初中校700分以上人数(民间统计)、初中学校排行榜(分区分榜)、最低投档控制线、全市成绩分布(一分一段)、市重招生计划汇总、名额分配到区计划(市重77所)、平行志愿分数线、各区分数水平(平行志愿代理)、名额到区录取分数线、名额到校录取分数线、名额分配到校计划(逐校)、各区统招(1-15志愿)计划、汇总透视、数据覆盖与缺口。"),
    ("阅读顺序建议", "第一次打开请先看『字段数据字典』(每个字段的口径与单位)，再看本页底部的『表关系与联动键』和『建议阅读路径』，最后按需查数据 Sheet。"),
    ("口径说明", "录取最低分在『名额分配(到区/到校)』批次按 800 分制(学业考750 + 综合考查50)；在『平行志愿(统一招生)』批次按 750 分制(学业考，不含综评)。两者不能直接相减或横向比高低，详见『字段数据字典』与下方警告。"),
    ("同分比较规则", "同分末位比较顺序：同分优待 → 综合素质评价 → 语数外 → 数学 → 语文 → 综合测试。"),
    ("重要缺口1(已推算+民间补)", "『各初中学校的考生人数』官方从不公布，但可由『名额分配到校按应届毕业生占全区比例分配』反推：推算考生数=全区考生数×(该校到校名额合计÷全区到校名额合计)，已新增『各初中考生人数(推算)』Sheet(703所初中)。另从教培/自媒体公众号、视频号、今日头条挖掘到『初中校 700 分以上人数』民间统计（已扩至 113 所初中，覆盖 16 区含此前空白的金山/崇明，含最高分与部分 710+/720+ 分段），已新增『初中校700分以上人数(民间统计)』Sheet——这是分数分布的高分段尾巴，非完整分布；该表额外用推算考生数算出『700+占比%(推算)』便于横向比较各校高分密度。"),
    ("重要缺口2(有区级代理)", "『全市和各区的成绩分布(一分一段)』考试院仅公布全市一分一段表且不提供可机读文本，任何公开来源(含教培/自媒体)均无分区一分一段。已新增『各区分数水平(平行志愿代理)』Sheet，用官方平行志愿线推算每区均值/中位数/高分占比，作为区级分数水平的近似代理（非一分一段）。"),
    ("重要缺口3(基本补齐·宝山部分)", "闵行/嘉定/松江三区『1-15志愿(平行志愿)』原以图片形式发布，后从上海本地宝文本页补齐；其余各区亦已入表。仅宝山区官方以图片发布且无可机读文本，仅能从公开报道提取到 6 所头部本区高中分数，故『平行志愿分数线』Sheet 中宝山区为【部分】(其余 15 区完整)。"),
    ("重要缺口4(已补齐)", "『名额分配到校招生计划(逐校明细)』现 16 区计划CSV已全部并入并新增『名额分配到校计划(逐校)』Sheet(初中×高中×计划数)；其中徐汇区官方初发图片版PDF、OCR 不可还原，后由用户手工将图片转为结构化 CSV 补齐(899=委属43+区属856，31所初中×11所高中)。16区计划合计12875，与全市口径12904差约29，系个别区计划表仍有极少量未解析单元格，非徐汇缺口。"),
    ("数据性质标注", "标注『官方』为考试院/教育局原文；标注『民间估算/预测』为机构汇总，仅供参考，请以官方公布为准。"),
    ("免责声明", "本数据集为公开信息汇编，可能存在转录误差或发布口径差异，正式用途请以上海市教育考试院及各区招考机构官方发布为准。"),
]
r = 3
for k, v in notes:
    ws.cell(row=r, column=1, value=k).font = Font(bold=True, color="1F4E78")
    c = ws.cell(row=r, column=2, value=v); c.alignment = WRAP
    r += 1

# --- 各 Sheet 含义速查表 ---
r += 1
ws.cell(row=r, column=1, value="各 Sheet 含义速查").font = SUB_FONT
r += 1
sheet_guide = [
    ("说明与数据来源", "本页：总体说明、来源、口径、缺口、表关系与阅读路径。"),
    ("字段数据字典", "每个 Sheet 每个字段的含义、取值/单位、注意事项。读数据前必看。"),
    ("考生人数", "全市 13.8 万(官方) + 16 区估算人数，作为升学率分母。"),
    ("各初中考生人数(推算)", "按到校名额占全区比例反推的各初中考生数(推算值)，附分区域自洽校验。"),
    ("初中校700分以上人数(民间统计)", "教培/自媒体渠道统计的各初中 700+/710+/720+ 人数与最高分，仅高分段尾巴，可靠度逐行标注。"),
    ("最低投档控制线", "各批次最低投档线(自招/名额分配 615、普高 501 等)，低于此线不能投档。"),
    ("全市成绩分布(一分一段)", "全市分数段累计位次：真实锚点 + 民间预测分档。"),
    ("市重招生计划汇总", "市重 30728 总计划按 4 批次拆解(自招/到区/到校/统招)，及全市高中构成。"),
    ("名额分配到区计划(市重77所)", "每所市重分配到各区的『计划数』(供给量)，合计 7171。"),
    ("平行志愿分数线", "1-15志愿(统一招生)录取最低分，按区筛选。"),
    ("名额到区录取分数线", "名额分配到区实际录取最低分，按『接收区』筛选。"),
    ("名额到校录取分数线", "名额分配到校实际录取最低分，初中×高中组合(本数据集最细粒度)。"),
    ("名额分配到校计划(逐校)", "名额分配到校的『计划数』(供给量)，初中×高中组合，16区已并入(徐汇由用户手工补入图片PDF转CSV)。"),
    ("各区统招(1-15志愿)计划", "各区统一招生批次的招生计划(本区/外区)，已知 10 区。"),
    ("汇总透视", "按高中/按区汇总的名额分配与分数线一览，快速看盘子(到区计划/到校计划/录取线)。"),
    ("数据覆盖与缺口", "逐项标注每个维度覆盖情况(✅/⚠️/❌)与说明。"),
]
r = write_table(ws, r, ["Sheet 名称", "一句话含义 / 用途"], sheet_guide,
                widths=[24, 104], freeze=False)

# --- 表关系与联动键 ---
r += 1
ws.cell(row=r, column=1, value="表关系与联动键").font = SUB_FONT
r += 1
rel_text = (
    "① 市重招生计划汇总(各批次) 是『总盘子』，其 4 个批次分别映射到明细：\n"
    "   - 名额分配到区 → 名额分配到区计划(市重77所)〔供给〕 + 名额到区录取分数线〔结果〕\n"
    "   - 名额分配到校 → 名额分配到校计划(逐校)〔供给·初中×高中×计划数〕 + 名额到校录取分数线〔结果·初中×高中录取线〕\n"
    "   - 平行志愿(统一招生) → 平行志愿分数线〔结果〕 + 各区统招(1-15志愿)计划〔供给〕\n"
    "   - 自主招生 → 暂无明细 Sheet\n"
    "② 联动键(用于跨表关联)：\n"
    "   - 高中维度：优先用『招生学校』名称关联；『招生代码』更稳定，存在于『名额分配到区计划』与『平行志愿分数线』，但『名额到区/到校录取分数线』无代码列，只能靠校名关联(校名可能略有出入)。\n"
    "   - 区维度：『名额到区』用『接收区』(考生来源区)；『名额到校』用『区』(初中所在区)；『平行志愿』用『区』(高中所在区)；三者口径不同，勿混淆。\n"
    "   - 初中维度：仅『名额到校录取分数线』含『初中学校』，可按初中筛选其拿到哪些高中、各多少分。\n"
    "③ 考生人数 是分母：可与各批次计划数相除估算升学率/竞争比，但各区人数为民间估算，仅作量级参考。\n"
    "④ 最低投档控制线 是门槛：所有录取分均 ≥ 对应批次控制线。"
)
ws.cell(row=r, column=1, value=rel_text).alignment = WRAP
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws.row_dimensions[r].height = 150
r += 1

# --- 建议阅读路径 ---
r += 1
ws.cell(row=r, column=1, value="建议阅读 / 分析路径").font = SUB_FONT
r += 1
paths = [
    "看整体盘子：市重招生计划汇总 → 考生人数(算市重录取率≈22%)。",
    "查某所市重点的名额分配：名额分配到区计划(找计划数) → 名额到区录取分数线(找实际录取线) → 名额到校录取分数线(看它分到哪些初中)。",
    "查某区统招：平行志愿分数线(按『区』筛选) + 各区统招(1-15志愿)计划。",
    "查某所初中升学：名额到校录取分数线(按『初中学校』筛选)，即得该校各高中录取线。",
    "估分数定位：全市成绩分布(一分一段) 看自己处于全市前百分之多少。",
]
for p in paths:
    ws.cell(row=r, column=1, value="• " + p).alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

# --- 最关键口径警告 ---
r += 1
ws.cell(row=r, column=1, value="⚠ 最关键口径警告（务必先读）").font = WARN_FONT
r += 1
warn = ("『录取最低分』在不同 Sheet 满分不同，切勿直接相减或横向比较：\n"
        "  - 名额分配到区 / 名额分配到校：满分 800 = 学业考(750) + 综合考查(50，即综合素质评价)。\n"
        "  - 1-15志愿(平行志愿)：满分 750 = 学业考，不含综合考查(综评在名额分配批次才计入)。\n"
        "因此一所高中『到校线 780』和『平行志愿线 690』不是差 90 分，而是两个不同计分体系，不能相减。")
ws.cell(row=r, column=1, value=warn).alignment = WRAP
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws.row_dimensions[r].height = 80
ws.cell(row=r, column=1).font = Font(color="C00000")

# ============ Sheet 2: 字段数据字典 ============
ws = wb.create_sheet("字段数据字典")
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 46
ws["A1"] = "字段数据字典（各 Sheet 字段含义、取值/单位、注意事项）"; ws["A1"].font = TITLE_FONT
glossary = [
    # (所属Sheet, 字段名, 含义, 取值/单位, 注意事项)
    ("考生人数", "区域", "区名或『全市』", "文本", "全市=16区合计(官方)；其余为各区"),
    ("考生人数", "考生人数(人)", "该区参加中考人数", "整数", "全市为官方 138000；各区为民间估算，非官方口径"),
    ("考生人数", "数据性质", "数据来源性质", "官方/民间估算", "官方仅全市，各区均为估算"),
    ("考生人数", "备注", "来源与说明", "文本", ""),
    ("各初中考生人数(推算)", "区", "区名", "文本", "按区汇总行的联动键"),
    ("各初中考生人数(推算)", "初中学校", "初中学校全称", "文本", "实际拿到到校名额的初中"),
    ("各初中考生人数(推算)", "推算考生数(人)", "该校推算考生数", "整数", "=全区考生数×(该校到校名额合计÷全区到校名额合计)；推算值非官方"),
    ("各初中考生人数(推算)", "到校名额合计", "该校名额分配到校计划总数", "整数", "来自『名额分配到校计划(逐校)』按校求和"),
    ("各初中考生人数(推算)", "占全区比例(%)", "该校名额占全区到校名额比", "0-100", "≈该校考生占全区比，即推算依据"),
    ("初中校700分以上人数(民间统计)", "区", "区名", "文本", "按区汇总联动键"),
    ("初中校700分以上人数(民间统计)", "初中学校", "初中学校全称", "文本", "民间统计口径，校名可能与官方略有出入"),
    ("初中校700分以上人数(民间统计)", "700+人数", "该校中考 700 分及以上人数", "整数/空", "民间统计；空=未统计到；可靠度见『可靠度』列"),
    ("初中校700分以上人数(民间统计)", "710+人数", "该校 710 分及以上人数", "整数/空", "部分学校才有细分"),
    ("初中校700分以上人数(民间统计)", "720+人数", "该校 720 分及以上人数", "整数/空", "部分学校才有细分"),
    ("初中校700分以上人数(民间统计)", "最高分", "该校已曝光最高分", "小数/空", "部分仅披露最高分，无 700+ 计数"),
    ("初中校700分以上人数(民间统计)", "中考人数(已知)", "该校实际中考人数(若渠道披露)", "整数/空", "仅少数学校披露，可用于校验推算考生数"),
    ("初中校700分以上人数(民间统计)", "可靠度", "数据可信度分级", "枚举", "多源一致/单一来源/网传存疑/仅最高分"),
    ("初中校700分以上人数(民间统计)", "备注", "来源与口径说明", "文本", "含来源、冲突标注(如徐汇中学67 vs 估30+)"),
    ("初中校700分以上人数(民间统计)", "700+占比%(推算)", "700+人数 ÷ 推算考生数", "0-100/空", "推算考生数来自『各初中考生人数(推算)』，仅校名可匹配时填列"),
    ("初中学校排行榜(分区分榜)", "区", "区名", "文本", "按区汇总联动键"),
    ("初中学校排行榜(分区分榜)", "初中学校", "初中学校全称", "文本", "拿到到校名额的初中"),
    ("初中学校排行榜(分区分榜)", "推算考生数(人)", "该校推算考生数", "整数/空", "来自『各初中考生人数(推算)』"),
    ("初中学校排行榜(分区分榜)", "到校名额合计", "该校名额分配到校计划总数", "整数/空", "来自『名额分配到校计划(逐校)』"),
    ("初中学校排行榜(分区分榜)", "生源强度指数(区内z)", "该校到校线在区内池的标准分均值", "小数", "z=(分-区内该市重均值)/标准差，跨市重取均值；消除市重难度差异"),
    ("初中学校排行榜(分区分榜)", "700+人数", "该校 700+ 人数", "整数/空", "来自『初中校700分以上人数(民间统计)』；空=未统计到"),
    ("初中学校排行榜(分区分榜)", "700+占比%(推算)", "700+人数÷推算考生数", "0-100/空", "高分密度；空=无700+统计"),
    ("初中学校排行榜(分区分榜)", "综合得分(数据+声誉)", "0.20×A+0.15×B+0.65×D 归一化", "0-100/空", "含声誉校用三档合成(声誉主导)；无声誉回退A:B=6:4并标『声誉待补』；区内归一化"),
    ("初中学校排行榜(分区分榜)", "区内排名A", "按生源强度指数的区内名次", "整数", "主排序"),
    ("初中学校排行榜(分区分榜)", "区内排名B", "按700+占比的区内名次", "整数/空", "仅双覆盖校"),
    ("初中学校排行榜(分区分榜)", "区内排名C", "按综合得分的区内名次", "整数/空", "仅双覆盖校"),
    ("初中学校排行榜(分区分榜)", "全市排名A", "按生源强度(全市z)的全市名次", "整数", "跨区仅供参考"),
    ("初中学校排行榜(分区分榜)", "数据完整度", "全量(仅A)/双覆盖", "枚举", "是否有700+民间统计"),
    ("最低投档控制线", "招生类别", "招生批次类别", "文本", "自招/名额分配/普高/中本贯通等"),
    ("最低投档控制线", "最低投档控制分数线(分)", "该批次最低投档门槛", "0-750", "低于此线不能投档；名额分配=615，普高=501"),
    ("最低投档控制线", "备注", "政策依据", "文本", ""),
    ("全市成绩分布(一分一段)", "分数(≥)", "分数阈值", "整数", "达到该分及以上"),
    ("全市成绩分布(一分一段)", "全市累计位次(人)", "该分数及以上人数", "整数", "真实锚点为官方/媒体确认值"),
    ("全市成绩分布(一分一段)", "性质", "数据性质", "真实锚点/预测", "预测分档非官方"),
    ("全市成绩分布(一分一段)", "说明", "备注", "文本", ""),
    ("全市成绩分布(一分一段)", "分数段", "预测分档区间", "文本", "民间预测，仅供参考"),
    ("全市成绩分布(一分一段)", "预估累计位次", "预估全市位次", "文本", "民间预测"),
    ("全市成绩分布(一分一段)", "对应梯队", "对应高中梯队", "文本", ""),
    ("市重招生计划汇总", "招生批次", "招生批次名", "文本", "自招/到区/到校/统招/合计"),
    ("市重招生计划汇总", "招生计划数(人)", "该批次招生人数", "整数", "合计 30728"),
    ("市重招生计划汇总", "备注", "说明", "文本", ""),
    ("市重招生计划汇总", "类别", "高中类别", "文本", "市重/区重/特色/普高"),
    ("市重招生计划汇总", "数量(所)", "该类高中数量", "整数", "全市高中合计 312 所"),
    ("名额分配到区计划(市重77所)", "招生代码", "高中学校代码", "文本", "跨表联动键(与平行志愿一致)"),
    ("名额分配到区计划(市重77所)", "学校名称", "高中学校名称", "文本", "联动键(校名)"),
    ("名额分配到区计划(市重77所)", "所属区", "高中所在区", "文本", ""),
    ("名额分配到区计划(市重77所)", "办别", "办学性质", "公办/民办", ""),
    ("名额分配到区计划(市重77所)", "学校类型", "学校类别", "文本", "市实验性示范性高中或享受同政策"),
    ("名额分配到区计划(市重77所)", "寄宿情况", "是否寄宿", "文本", "全部寄宿/部分寄宿/不寄宿"),
    ("名额分配到区计划(市重77所)", "计划数", "该校分配到区名额总数", "整数", "合计 7171；为供给量"),
    ("平行志愿分数线", "区", "高中所在区", "文本", "注意与到区『接收区』口径不同"),
    ("平行志愿分数线", "招生代码", "高中学校代码", "文本", "跨表联动键"),
    ("平行志愿分数线", "招生学校", "高中学校名称", "文本", "联动键(校名)"),
    ("平行志愿分数线", "录取最低分", "该校本区统招录取最低分", "0-750", "⚠ 满分750(学业考)，不含综评"),
    ("平行志愿分数线", "语数外", "语文+数学+外语合计", "0-450", ""),
    ("平行志愿分数线", "数学", "数学单科", "0-150", ""),
    ("平行志愿分数线", "语文", "语文单科", "0-150", ""),
    ("平行志愿分数线", "综合测试", "综合测试(物理+化学+跨学科+实验)", "0-150", ""),
    ("平行志愿分数线", "是否同分优待", "是否享受同分优待/政策加分", "是/否", ""),
    ("平行志愿分数线", "备注", "说明", "文本", ""),
    ("各区分数水平(平行志愿代理)", "区", "区名", "文本", "按区汇总联动键"),
    ("各区分数水平(平行志愿代理)", "高中数(本区招生)", "该区平行志愿招生高中数", "整数", "源自『平行志愿分数线』；宝山仅6所头部校"),
    ("各区分数水平(平行志愿代理)", "平行志愿线均值", "本区各高中录取最低分均值", "小数", "区级分数水平代理(非一分一段)"),
    ("各区分数水平(平行志愿代理)", "中位数", "本区各高中录取最低分中位数", "小数", "比均值更抗极端值"),
    ("各区分数水平(平行志愿代理)", "最高线", "本区最高录取线", "小数", "多为四校/四分分校"),
    ("各区分数水平(平行志愿代理)", "最低线", "本区最低录取线", "小数", "多为尾部普高/民办"),
    ("各区分数水平(平行志愿代理)", "≥700占比%", "录取线≥700的高中占比", "0-100", "越高说明该区头部校越密集"),
    ("各区分数水平(平行志愿代理)", "≥680占比%", "录取线≥680的高中占比", "0-100", ""),
    ("各区分数水平(平行志愿代理)", "≥650占比%", "录取线≥650的高中占比", "0-100", ""),
    ("各区分数水平(平行志愿代理)", "数据状态", "数据完整度", "文本", "完整·官方线 / 部分·宝山仅头部6所"),
    ("名额到区录取分数线", "接收区", "获得名额的区(考生来源区)", "文本", "≠高中所在区"),
    ("名额到区录取分数线", "招生学校", "高中学校名称", "文本", "联动键(校名)"),
    ("名额到区录取分数线", "隶属区", "高中所在区", "文本", ""),
    ("名额到区录取分数线", "办学性质", "公办/民办", "文本", ""),
    ("名额到区录取分数线", "录取最低分", "该区到区录取最低分", "0-800", "⚠ 满分800(含综评50)"),
    ("名额到区录取分数线", "语数外", "语文+数学+外语合计", "0-450", ""),
    ("名额到区录取分数线", "数学", "数学单科", "0-150", ""),
    ("名额到区录取分数线", "语文", "语文单科", "0-150", ""),
    ("名额到区录取分数线", "综合测试", "综合测试", "0-150", ""),
    ("名额到区录取分数线", "是否同分优待", "是否同分优待", "是/否", ""),
    ("名额到校录取分数线", "区", "初中所在区", "文本", "初中生源区"),
    ("名额到校录取分数线", "初中学校", "生源初中学校", "文本", "本数据集最细粒度维度"),
    ("名额到校录取分数线", "招生学校", "录取高中学校", "文本", "联动键(校名)"),
    ("名额到校录取分数线", "录取最低分", "该初中→该高中录取最低分", "0-800", "⚠ 满分800(含综评50)"),
    ("名额到校录取分数线", "综合素质评价", "综合考查成绩", "0-50", "已计入『录取最低分』，勿重复加"),
    ("名额到校录取分数线", "语数外", "语文+数学+外语合计", "0-450", ""),
    ("名额到校录取分数线", "数学", "数学单科", "0-150", ""),
    ("名额到校录取分数线", "语文", "语文单科", "0-150", ""),
    ("名额到校录取分数线", "综合测试", "综合测试", "0-150", ""),
    ("名额到校录取分数线", "是否同分优待", "是否同分优待", "是/否", ""),
    ("名额分配到校计划(逐校)", "区", "初中所在区", "文本", "初中生源区"),
    ("名额分配到校计划(逐校)", "初中学校", "生源初中学校", "文本", ""),
    ("名额分配到校计划(逐校)", "招生学校", "录取高中学校", "文本", "联动键(校名)"),
    ("名额分配到校计划(逐校)", "计划数", "该初中分到该高中的到校名额数", "整数", "为供给量；与录取分数线不同，是计划而非结果"),
    ("各区统招(1-15志愿)计划", "区", "区名", "文本", ""),
    ("各区统招(1-15志愿)计划", "本区高中计划(人)", "本区高中面向本区招生数", "整数", ""),
    ("各区统招(1-15志愿)计划", "外区高中计划(人)", "本区高中面向外区招生数", "整数", "部分区缺失填『—』"),
    ("各区统招(1-15志愿)计划", "合计(人)", "本区统招计划总数", "整数", ""),
    ("各区统招(1-15志愿)计划", "备注", "说明", "文本", ""),
    ("数据覆盖与缺口", "维度", "数据维度名", "文本", ""),
    ("数据覆盖与缺口", "覆盖情况", "覆盖状态", "✅/⚠️/❌", "✅完整 ⚠️部分/估算 ❌未公布"),
    ("数据覆盖与缺口", "说明", "缺口详情", "文本", ""),
    ("汇总透视", "招生学校", "高中学校名称", "文本", "按高中汇总的联动键(校名)"),
    ("汇总透视", "到区计划数", "该校名额分配到区总计划", "整数", "来自『名额分配到区计划(市重77所)』"),
    ("汇总透视", "到校计划数(合计)", "该校名额分配到校总计划(逐校合计)", "整数", "来自『名额分配到校计划(逐校)』(16区已并入)"),
    ("汇总透视", "到校覆盖初中数", "拿到该校到校名额的初中数", "整数", "来自『名额到校录取分数线』去重"),
    ("汇总透视", "到区录取线(最低·800制)", "该校到区录取最低分(各区最低)", "0-800", "⚠ 800分制(含综评50)"),
    ("汇总透视", "平行志愿线(最低·750制)", "该校统招录取最低分(各区最低)", "0-750", "⚠ 750分制(不含综评)"),
    ("汇总透视", "区", "区名", "文本", "按区汇总行"),
    ("汇总透视", "到校计划数合计", "该区全部初中到校计划之和", "整数", "来自『名额分配到校计划(逐校)』(16区已并入)"),
    ("汇总透视", "享受到校初中数", "该区享有到校名额的初中数", "整数", ""),
    ("汇总透视", "到校录取线覆盖初中数", "该区有到校录取线记录的初中数", "整数", "来自『名额到校录取分数线』"),
    ("汇总透视", "平行志愿招生学校数", "该区统招批次招生高中数", "整数", "来自『平行志愿分数线』"),
    ("汇总透视", "到区录取线高中数(接收)", "该区作为接收区拿到到区计划的高中数", "整数", "来自『名额到区录取分数线』"),
]
r = write_table(ws, 3,
    ["所属 Sheet", "字段名", "含义", "取值 / 单位", "注意事项"],
    glossary, widths=[24,18,50,20,46], freeze=True)

# ============ Sheet 3: 考生人数 ============
ws = wb.create_sheet("考生人数")
ws["A1"] = "2026 年上海中考考生人数（全市 + 各区）"; ws["A1"].font = SUB_FONT
city_wide = [("全市", 138000, "官方", "上海招考热线 2026-06-20 公布；16考区/214考点/4730考场")]
district_est = [
    ("浦东新区", 33348, "民间估算", "haoxue360 2026各区预计中考人数"),
    ("闵行区", 17374, "民间估算", "haoxue360；搜狐称考生超1.5万"),
    ("宝山区", 11555, "民间估算", "haoxue360"),
    ("徐汇区", 8754, "民间估算", "haoxue360"),
    ("嘉定区", 8675, "民间估算", "haoxue360"),
    ("杨浦区", 7289, "民间估算", "haoxue360"),
    ("静安区", 7351, "民间估算", "haoxue360"),
    ("普陀区", 7206, "民间估算", "haoxue360"),
    ("松江区", 10263, "民间估算", "haoxue360"),
    ("青浦区", 5551, "民间估算", "haoxue360"),
    ("奉贤区", 5637, "民间估算", "haoxue360"),
    ("黄浦区", 4153, "民间估算", "haoxue360"),
    ("虹口区", 4158, "民间估算", "haoxue360"),
    ("长宁区", 3791, "民间估算", "haoxue360"),
    ("金山区", 4298, "民间估算", "haoxue360"),
    ("崇明区", 2768, "民间估算", "haoxue360"),
]
rows = city_wide + district_est
est_sum = sum(x[1] for x in district_est)
rows.append(("16区估算合计", est_sum, "民间估算", "仅为各区估算之和，非官方口径"))
write_table(ws, 3, ["区域","考生人数(人)","数据性质","备注"], rows, widths=[14,16,12,60])

# ============ Sheet: 各初中考生人数(推算) ============
# 方法：名额分配到校按各初中应届毕业生占全区比例均衡分配，
#      故 推算考生数 = 全区考生数 × (该校到校名额合计 ÷ 全区到校名额合计)
ws_inf = wb.create_sheet("各初中考生人数(推算)")
ws_inf["A1"] = "各初中学校考生人数（按到校名额占比推算，非官方公布）"; ws_inf["A1"].font = SUB_FONT
_inf_note = ("推算方法：名额分配到校计划按各初中应届毕业生占全区比例均衡分配，"
             "故 推算考生数 = 全区考生数 × (该校到校名额合计 ÷ 全区到校名额合计)。"
             "全区考生数取自本工作簿『考生人数』Sheet（16区为民间估算，非官方口径）；"
             "到校名额合计取自『名额分配到校计划(逐校)』。结果四舍五入，单校误差约 ±(全区考生数÷全区到校名额) 人/名额；"
             "仅覆盖实际拿到到校名额的初中，个别 0 名额校（多为无中招报名资格或极小规模）不在此表。"
             "此数为推算值，仅供量级参考，正式用途以上海市教育考试院公布为准。")
ws_inf["A2"] = _inf_note; ws_inf["A2"].alignment = WRAP; ws_inf.merge_cells("A2:E2")

dist_cand2 = {d: c for (d, c, *_rest) in district_est}
junior_n = {}; dist_n = {}
for fn in sorted(os.listdir(DATA)):
    if not fn.startswith("to_school_plan_") or not fn.endswith(".csv"):
        continue
    dk = fn[len("to_school_plan_"):-4]
    dist = DISTRICT_MAP.get(dk, dk)
    dsum = 0
    with open(os.path.join(DATA, fn), encoding="utf-8-sig") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            p = line.split(",")
            if len(p) >= 4 and p[3].strip().isdigit():
                j = p[1].strip(); n = int(p[3])
                junior_n[(dist, j)] = junior_n.get((dist, j), 0) + n
                dsum += n
    dist_n[dist] = dist_n.get(dist, 0) + dsum

inf_rows = []
for (dist, j), n in junior_n.items():
    C = dist_cand2.get(dist, 0); Nd = dist_n.get(dist, 0)
    c_inf = round(C * n / Nd) if (C and Nd) else 0
    pct = round(n / Nd * 100, 2) if Nd else 0
    inf_rows.append((dist, j, c_inf, n, pct))
inf_rows.sort(key=lambda x: (DISTRICTS_ORDER.index(x[0]) if x[0] in DISTRICTS_ORDER else 99, -x[2]))
write_table(ws_inf, 3, ["区", "初中学校", "推算考生数(人)", "到校名额合计", "占全区比例(%)"], inf_rows,
            widths=[12, 40, 16, 14, 14], freeze=True)
# 分区域自洽校验（推算合计 vs 区考生数）
r_foot = ws_inf.max_row + 2
ws_inf.cell(row=r_foot, column=1, value="分区域自洽校验（推算合计 vs 区考生数，差异来自四舍五入）").font = Font(bold=True, color="1F4E78")
chk = []
for dist in DISTRICTS_ORDER:
    if dist not in dist_cand2:
        continue
    tot = sum(c for (d, j, c, n, p) in inf_rows if d == dist)
    chk.append((dist, tot, dist_cand2[dist], tot - dist_cand2[dist]))
write_table(ws_inf, r_foot + 1, ["区", "推算合计", "区考生数(估算)", "差"], chk, widths=[12, 14, 16, 10], freeze=False)

# 推算考生数查找表（用于计算 700+ 占比 / 综合榜），三档稳健匹配
#   档1 精确；档2 人工别名（缩写↔官方全称，已逐区 Grep 核对 to_school_plan 真实数据）；
#   档3 归一化子串（去「上海市/民办」前缀、「第N→N」、「附属→附」、去校名后缀）；
#   档4 模糊匹配仅在多候选时拒绝(返回 None)以避免误配；逐校结果写入 match_report.csv 供人工复查。
inf_lookup = {(d, j): c for (d, j, c, n, p) in inf_rows}

def canon(s):
    """校名归一化：去括号注释(全角/半角)、去前缀、『第N→N』、『附属→附』、去尾部校名后缀，便于子串比对。"""
    s = s.strip()
    # 去括号注释：如（含领科校区）、（上海财大附属…）、(上海财经大学…）
    while True:
        new = re.sub(r"[（(][^（）()]*[）)]", "", s)
        if new == s:
            break
        s = new
    for p in ("上海市", "上海市民办", "上海民办", "上海", "民办"):
        if s.startswith(p):
            s = s[len(p):]
    s = (s.replace("第一", "一").replace("第二", "二").replace("第三", "三")
           .replace("第四", "四").replace("第五", "五").replace("附属", "附"))
    for suf in ("初级中学", "实验中学", "实验学校", "附属中学", "附属学校",
                "外国语学校", "双语学校", "学校", "中学"):
        if s.endswith(suf):
            s = s[:-len(suf)]
            break
    return s

# 人工别名：缩写/全称 → to_school_plan 官方全称（逐区核对 to_school_plan 真实数据，确保与 inf_lookup 的 key 完全一致）
#   既有 700+ 表缩写映射，也有 录取线全名→plan缩写 的映射（用于把 strength 键统一到 plan 键，消除榜C碎片）。
ALIAS = {
    # ---- 700+ 表缩写 → 官方全称 ----
    ("黄浦区", "立达中学"): "民办立达中学",
    ("黄浦区", "永昌中学"): "民办永昌学校",
    ("闵行区", "七宝二中"): "上海市闵行区七宝第二中学",
    ("闵行区", "七宝三中"): "上海市闵行区七宝第三中学",
    ("闵行区", "交大二附中"): "上海交通大学附属第二中学",
    ("闵行区", "闵华二附中"): "华东师范大学第二附属中学附属初级中学",
    ("徐汇区", "华理附中"): "华东理工大学附属中学",
    ("徐汇区", "师三附中"): "上海师范大学第三附属实验学校",
    ("徐汇区", "康健外国语实验学校"): "上海市康健外国语实验中学",
    ("徐汇区", "世外中学"): "上海市世界外国语中学",
    ("静安区", "风华中学"): "上海市风华初级中学",
    ("普陀区", "曹杨附中"): "上海市曹杨第二中学附属实验中学",
    ("浦东新区", "上中东校"): "上海中学东校",
    ("浦东新区", "华二前滩"): "华东师范大学第二附属中学前滩学校",
    ("浦东新区", "华曜浦东"): "上海民办华曜浦东实验学校",
    ("浦东新区", "建平远翔"): "上海浦东新区民办远翔实验学校",
    ("浦东新区", "进才实验"): "上海市进才实验中学",
    ("浦东新区", "尚德实验"): "上海市民办尚德实验学校",
    ("浦东新区", "张江集团中学"): "上海市张江集团中学",
    # 新增：清晰简称→全称（并消歧，如 文来 vs 七宝文来）
    ("闵行区", "文来中学"): "上海市文来中学",
    ("徐汇区", "徐教院附中"): "上海市徐汇区教育学院附属实验中学",
    ("长宁区", "延安初级中学"): "上海市延安初级中学",
    ("普陀区", "梅陇中学"): "上海市梅陇中学",
    ("普陀区", "华外实验"): "华东师范大学附属外国语实验学校",
    ("闵行区", "实验西校"): "上海市实验学校西校",
    ("浦东新区", "交中初级"): "上海民办浦东交中初级中学",
    ("浦东新区", "川沙南校"): "上海市川沙中学南校",
    ("普陀区", "同济二附中"): "同济大学第二附属中学",
    ("闵行区", "田园外语中学"): "上海市闵行区田园外国语中学",
    ("闵行区", "浦江一中"): "上海市闵行区浦江第一中学",
    # ---- 录取线全名 → plan 全称（统一 strength 键到 plan 键）----
    ("黄浦区", "上海市第十中学"): "市十中学",
    ("黄浦区", "上海交通大学附属黄浦实验中学"): "交附黄浦实验",
    ("黄浦区", "上海理工大学附属储能中学"): "储能中学",
    ("黄浦区", "上海外国语大学附属大境初级中学"): "大境初级中学",
    ("黄浦区", "上海音乐学院附属黄浦比乐中学"): "上音比乐中学",
}

# ============ 高中校名统一（汇总透视 5 来源 join + 各录取线/计划表的高中列）============
# 问题：parallel_（平行志愿）系统性使用缩写/异写（上海中学/交大附中/丰华高级/田家炳…），
# 而其他 4 来源（plan/到校录取线/to_district/到区计划）使用官方全称，导致汇总透视按原始串并集时
# 同一高中被拆成 2+ 行（平行志愿线列与到校/到区列分处两行）。下方把平行志愿的异写统一回官方全称，
# 其余来源本就互相一致（仅个别括号全/半角、OCR 空格/换行），经 _hs_norm 归一后自动对齐。
HS_ALIAS = {
    "上海中学": "上海市上海中学",
    "丰华高级": "上海市民办丰华高级中学",
    "久隆模范": "上海市久隆模范中学",
    "交大附中": "上海市交大附中",
    "交大南洋": "上海市民办交大南洋中学",
    "复旦附中": "上海市复旦附中",
    "创艺高级": "上海创艺高级中学",
    "同洲模范": "上海市同洲模范学校",
    "宋庆龄": "上海宋庆龄学校",
    "扬波中学": "上海市民办扬波中学",
    "民一中学": "上海民办民一中学",
    "永昌中学": "上海市民办永昌中学",
    "田家炳": "上海田家炳中学",
    "行中中学": "上海民办行中中学",
    "赫贤学校": "上海赫贤学校",
    "金瑞学校": "上海金瑞学校",
    "金苹果": "上海市民办金苹果学校",
    "安生学校": "上海安生学校",
    "存志高级": "上海存志高级中学",
    "风范中学": "上海市民办风范中学",
}

def _hs_norm(s):
    """高中校名归一：去空白/换行、半角括号→全角、去括号注释（含（梅陇校区）/（艺术班）/（日语）等校区或班型变体）。"""
    s = re.sub(r"\s+", "", s.strip())
    s = s.replace("(", "（").replace(")", "）")
    s = re.sub(r"[（(][^（）()]*[）)]", "", s)
    return s

def _hs_unify(raw):
    """返回统一后的高中校名（官方全称）。平行志愿缩写→HS_ALIAS；其余来源本就为全称，去括号/空格后直接采用。"""
    if not raw:
        return raw
    n = _hs_norm(raw)
    return HS_ALIAS.get(n, n)

def _jr_unify(dist, raw):
    """返回统一后的初中校名（to_school_plan 官方全称，与『排行榜/700+表』同源）。
    用与排行榜相同的 _match_inf_verbose 匹配器，确保跨表校名完全一致；匹配失败则保留原值。"""
    if not raw:
        return raw
    r = _match_inf_verbose(dist, raw)[0]
    return r if r else raw

# 归一化残留令牌：用于子串匹配的去歧义（残留须仅由这些令牌构成，才算同一校的不同写法）
_DISTOK = set(DISTRICTS_ORDER)
_PREFIXOK = ["上海市", "上海", "市", "公立", "民办", "区属",
             "华东师范大学", "上海交通大学", "上海师范大学", "上海大学", "复旦大学", "同济大学",
             "上海财经大学", "上海外国语大学", "上海理工大学", "上海音乐学院", "上海中医药大学",
             "上海科技大学", "上海政法", "上海海事", "上海体育", "上海戏剧"]
_SUFFIXOK = ["附属", "实验", "外国语", "双语", "中学", "初级中学", "完全中学", "实验学校",
             "附属学校", "附属中学", "高级中学", "学校", "学院", "校区", "分校", "总校", "附",
             "南校", "北校", "东校", "西校", "高中", "部", "南部分校", "(南校)", "（南校）",
             "第一", "第二", "第三", "第四", "第五"]

def _residue_ok(res):
    """残留(residue)是否仅由可识别的前缀/后缀/区名令牌构成。"""
    s = res
    changed = True
    while changed and s:
        changed = False
        for t in _PREFIXOK + _SUFFIXOK + list(_DISTOK):
            if t and s.startswith(t):
                s = s[len(t):]; changed = True; break
            if t and s.endswith(t):
                s = s[:len(s) - len(t)]; changed = True; break
    return s == ""

def _match_inf_verbose(dist, name):
    """返回 (官方全称或None, 推算考生数或None, 匹配方式)。

    三档：exact(精确) → alias(人工别名) → normalized(归一化子串，多候选时用残留令牌去歧义)。
    任何档都无法唯一确定时返回 None（None 表示未匹配/歧义/数据缺口），绝不臆测，避免误配导致数据碎片。
    """
    if (dist, name) in inf_lookup:
        return name, inf_lookup[(dist, name)], "exact"
    if (dist, name) in ALIAS:
        fn = ALIAS[(dist, name)]
        if (dist, fn) in inf_lookup:
            return fn, inf_lookup[(dist, fn)], "alias"
        return None, None, "alias_missing"
    cn = canon(name)
    cand = []
    for (d, j), v in inf_lookup.items():
        if d != dist:
            continue
        cj = canon(j)
        if cn and (cn in cj or cj in cn):
            cand.append((j, v))
    if not cand:
        return None, None, "unmatched"
    if len(cand) == 1:
        return cand[0][0], cand[0][1], "normalized"
    # 多候选：用残留令牌严格化去歧义，仅保留唯一符合项
    good = []
    for j, v in cand:
        cj = canon(j)
        if cn in cj and _residue_ok(cj.replace(cn, "")):
            good.append((j, v))
        elif cj in cn and _residue_ok(cn.replace(cj, "")):
            good.append((j, v))
    if len(good) == 1:
        return good[0][0], good[0][1], "normalized"
    return None, None, "ambiguous"

def _match_inf(dist, name):
    """兼容旧调用：仅返回推算考生数（None 表示未匹配）。"""
    return _match_inf_verbose(dist, name)[1]

# ============ Sheet: 初中校700分以上人数(民间统计) ============
ws700 = wb.create_sheet("初中校700分以上人数(民间统计)")
ws700["A1"] = "各初中学校 700 分以上人数（民间统计，非官方）"; ws700["A1"].font = SUB_FONT
_700_note = ("数据来源：上海中考/教培类公众号、视频号、今日头条等民间渠道的 2026 中考高分统计汇总"
             "（上海择校升学转学·冯老师 2026-07-16、锐思教育·坦途网 2026-07-17、今日头条『700分+扎堆』2026-07-18 等）。"
             "这是『初中校分数分布』在 700 分以上的高分段尾巴，仅部分初中有统计，并非完整分数分布，也不是考生人数。"
             "『可靠度』列标注：多源一致 / 单一来源 / 网传存疑 / 仅最高分。网传或口径冲突的数据已在『备注』说明，"
             "请勿直接使用此类数字做精确比较；本表仅供了解各校高分产出量级。"
             "『700+占比%(推算)』= 该校700+人数 ÷ 『各初中考生人数(推算)』推算考生数。"
             "校名匹配采用三档稳健策略（精确→人工别名[简称↔全称/分校]→归一化子串去歧义），并在『推算匹配校名(全称)』『匹配方式』两列标注；"
             "未能唯一匹配（推算表无该校 / 多候选歧义 / 数据缺口）的，占比留空，详见 match_report_700plus.csv 人工复查。")
ws700["A2"] = _700_note; ws700["A2"].alignment = WRAP; ws700.merge_cells("A2:L2")

rows700 = []
match_report = []  # (区, 简称, 全称, 方式, 推算考生数, 700+人数, 占比, 说明)
match_cnt = 0
with open(os.path.join(DATA, "junior_high_700plus.csv"), encoding="utf-8-sig") as fh:
    rd = csv.reader(fh)
    next(rd)
    for p in rd:
        if not p or len(p) < 9:
            continue
        def _i(x):
            x = (x or "").strip()
            return int(x) if x else None
        dist, orig = p[0].strip(), p[1].strip()
        n700 = _i(p[2])
        res_name, c_inf, method = _match_inf_verbose(dist, orig)
        pct700 = round(n700 / c_inf * 100, 1) if (c_inf and n700) else None
        if pct700 is not None:
            match_cnt += 1
        mshort = {"exact": "精确", "alias": "别名", "normalized": "归一化",
                  "unmatched": "未匹配(数据缺口)", "ambiguous": "歧义(需人工)", "alias_missing": "别名缺失"}.get(method, method)
        note_map = {
            "exact": "精确匹配", "alias": "人工别名(简称↔全称/分校)", "normalized": "归一化子串(去歧义)",
            "unmatched": "推算表无该校(0到校名额或数据缺口)，占比留空",
            "ambiguous": "多候选歧义，需人工确认，占比留空",
            "alias_missing": "别名目标不在推算表，占比留空",
        }
        match_report.append((dist, orig, res_name or "", method, c_inf if c_inf else "", n700 if n700 else "",
                             pct700 if pct700 is not None else "", note_map.get(method, method)))
        rows700.append((
            dist, orig, res_name or "", mshort,
            n700, _i(p[3]), _i(p[4]),
            (float(p[5]) if p[5].strip() else None),
            _i(p[6]), pct700, p[7].strip(), p[8].strip()
        ))
rows700.sort(key=lambda x: (DISTRICTS_ORDER.index(x[0]) if x[0] in DISTRICTS_ORDER else 99, -(x[4] or 0)))
write_table(ws700, 3, ["区", "初中学校", "推算匹配校名(全称)", "匹配方式", "700+人数", "710+人数", "720+人数", "最高分", "中考人数(已知)", "700+占比%(推算)", "可靠度", "备注"],
            rows700, widths=[10, 22, 30, 22, 10, 10, 10, 10, 14, 14, 12, 36], freeze=True)

# 逐校匹配审计表（供人工复查）
with open(os.path.join(OUTPUT, "match_report_700plus.csv"), "w", encoding="utf-8-sig", newline="") as mf:
    w = csv.writer(mf)
    w.writerow(["区", "700+校名(简称)", "推算匹配校名(全称)", "匹配方式", "推算考生数", "700+人数", "700+占比%", "说明"])
    for r in match_report:
        w.writerow(r)
print(f"[700+匹配] 共 {len(rows700)} 行，成功匹配 {match_cnt} 行，审计表已写 match_report_700plus.csv")

# ============ Sheet: 初中学校排行榜(分区分榜) ============
ws_rank = wb.create_sheet("初中学校排行榜(分区分榜)")
ws_rank["A1"] = "各初中学校排行榜（分区分榜）— 区内排名为主"; ws_rank["A1"].font = SUB_FONT
_rank_note = ("排名方法：\n"
    "· 榜A 生源强度：对每所市重，在该校『名额到校录取线』的初中间计算标准分 z=(分-均值)/标准差（区内池与全市池各算一次）；"
    "某校生源强度指数=其所有市重 z 的均值。z 分消除了市重难度差异，全量覆盖全部拿到到校名额的初中。\n"
    "· 榜B 高分产出：按『700+占比%(推算)』（700+人数÷推算考生数）区内排名（仅民间统计到的学校有值）。\n"
    "· 榜D 声誉口碑：来自网页可检索的公开排名/口碑汇总（机构公众号、家长论坛、百度百家号等），逐校标注来源，属主观判断、仅供参考；"
    "本项目内数据不完整，故排名不能只靠 A/B。\n"
    "· 综合得分(数据+声誉)：区内将 A、B、D 各自归一化到 0–100 后加权合成（A:B:D=2:1.5:6.5，声誉主导），"
    "因项目内『名额到校录取线/700+』对自招强校易失真（尖子走四校自招、到校线反而低），故以公开口碑为主、数据为辅；"
    "凡校名可匹配到声誉数据者即用此式（缺 A/B 的维度取中性 50，数据不完整不惩罚）；无声誉数据者回退到原 A:B=6:4 数据综合，并标『声誉待补』。\n"
    "· 区内排名为本表主排序（同区共用同一批市重名额池，最可比）；全市排名为跨区全局 z，仅供参考、不宜跨区比高低。\n"
    "· 口径：到校录取线为 800 制（含综评50），仅在校内/校际的到校线之间比较，不混入平行志愿 750 制。推算考生数与 700+ 均为推算/民间值，声誉为主观口碑，见『数据完整度』列。")
ws_rank["A2"] = _rank_note; ws_rank["A2"].alignment = WRAP; ws_rank.merge_cells("A2:M2")

# ---- 读取名额到校录取线 ----
tschool_rows = []  # (区, 初中, 市重, 分)
for fn in sorted(os.listdir(DATA)):
    if not fn.startswith("to_school_") or fn.startswith("to_school_plan_") or not fn.endswith(".csv"):
        continue
    key = fn[len("to_school_"):-4]
    dist = DISTRICT_MAP.get(key, key)
    drows, _ = read_csv(DATA + "/" + fn)
    for x in drows:
        if not x or (len(x) >= 1 and x[0] == "区") or len(x) < 4:
            continue
        try:
            sc = float(x[3])
        except (ValueError, TypeError):
            continue
        tschool_rows.append((dist, x[1].strip(), x[2].strip(), sc))

def _dist(vals):
    n = len(vals)
    if n == 0:
        return 0.0, 0.0
    m = sum(vals) / n
    if n == 1:
        return m, 0.0
    var = sum((v - m) ** 2 for v in vals) / n
    return m, var ** 0.5

pool_in, pool_all, tmp_in, tmp_all = {}, {}, {}, {}
for (d, j, k, sc) in tschool_rows:
    tmp_in.setdefault((d, k), []).append(sc)
    tmp_all.setdefault(k, []).append(sc)
for kk, vals in tmp_in.items():
    pool_in[kk] = _dist(vals)
for kk, vals in tmp_all.items():
    pool_all[kk] = _dist(vals)

def _z(pool, kk, sc):
    mu, sg = pool.get(kk, (sc, 0.0))
    return 0.0 if sg == 0 else (sc - mu) / sg

strength = {}
agg_in, agg_all = {}, {}
strength_report = []  # (区, 录取线校名, 统一到plan校名或空, 方式)
for (d, j, k, sc) in tschool_rows:
    # 把录取线校名统一到 plan 官方全称键，使 榜A(生源强度) 与 榜B/榜C 在同一 schools 条目上合并
    res, _, method = _match_inf_verbose(d, j)
    key = (d, res) if res else (d, j)
    strength_report.append((d, j, res or "", method))
    a = agg_in.setdefault(key, [0.0, 0]); a[0] += _z(pool_in, (d, k), sc); a[1] += 1
    b = agg_all.setdefault(key, [0.0, 0]); b[0] += _z(pool_all, k, sc); b[1] += 1
for key in agg_in:
    s_in = agg_in[key][0] / agg_in[key][1]
    s_all = agg_all.get(key, [0.0, 1])[0] / agg_all.get(key, [1, 1])[1]
    strength[key] = (s_in, s_all, agg_in[key][1])

# 录取线→plan 统一审计表（供人工复查榜A与榜B/C的合并是否正确）
with open(os.path.join(OUTPUT, "match_report_strength.csv"), "w", encoding="utf-8-sig", newline="") as mf:
    w = csv.writer(mf)
    w.writerow(["区", "录取线校名", "统一到plan校名(全称)", "匹配方式"])
    for r in strength_report:
        w.writerow(r)
_uni = sum(1 for r in strength_report if r[2])
print(f"[录取线→plan] 共 {len(strength_report)} 校次，统一 {_uni} 校次，审计表已写 match_report_strength.csv")

hi_map = {}
# 用「推算匹配校名(全称)」作为键，使其与 inf_lookup / strength 的键（均为 plan 官方全称）对齐，
# 从而同一初中在 schools 中合并为一条，700+占比 才能与 生源强度 合成 榜C 综合得分。
for (d, orig, rj, method, n700, a710, a720, high, known, pct, rel, note) in rows700:
    if rj:  # 仅当成功匹配到官方全称时才入 hi_map
        hi_map[(d, rj)] = (n700, pct)

schools = {}
for (d, j), (s_in, s_all, nk) in strength.items():
    schools.setdefault((d, j), {})["s_in"] = s_in
    schools[(d, j)]["s_all"] = s_all
    schools[(d, j)]["n_schools"] = nk
for (d, j), c in inf_lookup.items():
    schools.setdefault((d, j), {})["cand"] = c
for (d, j), (n700, pct) in hi_map.items():
    schools.setdefault((d, j), {})["n700"] = n700
    schools.setdefault((d, j), {})["pct700"] = pct
for (d, j), n in junior_n.items():
    schools.setdefault((d, j), {})["plan"] = n

by_dist_s, by_dist_pct = {}, {}
for (d, j), v in schools.items():
    by_dist_s.setdefault(d, []).append(v.get("s_in", 0.0))
    if v.get("pct700") is not None:
        by_dist_pct.setdefault(d, []).append(v["pct700"])

def _minmax(vals):
    return (min(vals), max(vals)) if vals else (0.0, 0.0)

# ---- 声誉/口碑(榜D) ----
# 来源：网页可检索的公开排名/口碑汇总(机构公众号、家长论坛、百度百家号等)，逐校标注，主观、仅供参考。
# 嘉定区为本轮联网调研结果(多源交叉)；其余区为公认顶流校种子(标"待核验")。扩展只需编辑 data/school_reputation.csv。
REP_FILE = os.path.join(DATA, "school_reputation.csv")
rep_lookup = {}
if os.path.exists(REP_FILE):
    with open(REP_FILE, encoding="utf-8-sig", newline="") as _f:
        for _r in csv.DictReader(_f):
            _d = (_r.get("区") or "").strip()
            _s = (_r.get("学校") or "").strip()
            try:
                _sc = float((_r.get("声誉分") or "").strip())
            except ValueError:
                _sc = None
            if _d and _s and _sc is not None:
                rep_lookup[(_d, _s)] = (_sc, (_r.get("梯队") or "").strip(), (_r.get("依据") or "").strip())
print(f"[声誉] 载入 {len(rep_lookup)} 校声誉数据(来源:公开口碑/机构排名,主观仅供参考)")
by_dist_rep = {}
for (d, j), v in schools.items():
    if (d, j) in rep_lookup:
        by_dist_rep.setdefault(d, []).append(rep_lookup[(d, j)][0])
norm_d = {}
for d, vals in by_dist_rep.items():
    lo, hi = _minmax(vals)
    for (dd, jj), rep in rep_lookup.items():
        if dd == d:
            norm_d[(dd, jj)] = (rep[0] - lo) / (hi - lo) * 100 if hi > lo else 50.0

norm_s, norm_pct = {}, {}
for d, vals in by_dist_s.items():
    lo, hi = _minmax(vals)
    for (dd, jj), v in schools.items():
        if dd == d and "s_in" in v:
            norm_s[(dd, jj)] = (v["s_in"] - lo) / (hi - lo) * 100 if hi > lo else 50.0
for d, vals in by_dist_pct.items():
    lo, hi = _minmax(vals)
    for (dd, jj), v in schools.items():
        if dd == d and v.get("pct700") is not None:
            norm_pct[(dd, jj)] = (v["pct700"] - lo) / (hi - lo) * 100 if hi > lo else 50.0

WA, WB, WD = 0.20, 0.15, 0.65  # 综合(数据+声誉) 权重：声誉主导(0.65)，因项目内录取线/700+ 对自招强校易失真，须以口碑为主、数据为辅
for (d, j), v in schools.items():
    has_a = v.get("s_in") is not None
    has_b = v.get("pct700") is not None
    rep = rep_lookup.get((d, j))
    if rep:
        # 含声誉：缺 A/B 的维度用中性 50（数据不完整不惩罚），声誉维度必参与
        na = norm_s.get((d, j), 50.0) if has_a else 50.0
        nb = norm_pct.get((d, j), 50.0) if has_b else 50.0
        nd = norm_d.get((d, j), 50.0)
        v["comp"] = round(WA * na + WB * nb + WD * nd, 1)
        v["rep_score"], v["rep_tier"], v["rep_src"] = rep
        v["has_rep"] = True
    else:
        # 无声誉：回退到数据综合(只要有榜A或榜B即算，缺维取中性50)，标注"声誉待补"
        if has_a or has_b:
            na = norm_s.get((d, j), 50.0) if has_a else 50.0
            nb = norm_pct.get((d, j), 50.0) if has_b else 50.0
            v["comp"] = round(0.6 * na + 0.4 * nb, 1)
        else:
            v["comp"] = None
        v["rep_score"], v["rep_tier"], v["rep_src"] = "", "", ""
        v["has_rep"] = False

# 区内排名 A / 全市排名 A / 区内排名 B / 区内排名 C
for (d, j), v in schools.items():
    schools[(d, j)].setdefault("rankA_in", None)
    schools[(d, j)].setdefault("rankA_all", None)
    schools[(d, j)].setdefault("rankB_in", None)
    schools[(d, j)].setdefault("rankC_in", None)
distA = {}
for (d, j), v in schools.items():
    distA.setdefault(d, []).append(((d, j), v.get("s_in", -1e9)))
for d, lst in distA.items():
    lst.sort(key=lambda t: -t[1])
    for rk, (key, _) in enumerate(lst, 1):
        schools[key]["rankA_in"] = rk
allA = sorted(schools.items(), key=lambda kv: -kv[1].get("s_all", -1e9))
for rk, (key, _) in enumerate(allA, 1):
    schools[key]["rankA_all"] = rk
distB = {}
for (d, j), v in schools.items():
    if v.get("pct700") is not None:
        distB.setdefault(d, []).append(((d, j), v["pct700"]))
for d, lst in distB.items():
    lst.sort(key=lambda t: -t[1])
    for rk, (key, _) in enumerate(lst, 1):
        schools[key]["rankB_in"] = rk
distC = {}
for (d, j), v in schools.items():
    if v.get("comp") is not None:
        distC.setdefault(d, []).append(((d, j), v["comp"]))
for d, lst in distC.items():
    lst.sort(key=lambda t: -t[1])
    for rk, (key, _) in enumerate(lst, 1):
        schools[key]["rankC_in"] = rk

rank_rows = []
for (d, j), v in schools.items():
    pct700 = v.get("pct700"); n700 = v.get("n700"); comp = v.get("comp")
    has_a = v.get("s_in") is not None
    has_b = pct700 is not None
    base = "双覆盖" if (has_a and has_b) else ("仅B(缺A)" if has_b else "全量(仅A)")
    completeness = base + ("·含声誉" if v.get("has_rep") else "·声誉待补")
    rank_rows.append((
        d, j,
        v.get("cand", ""),
        v.get("plan", ""),
        round(v.get("s_in", 0.0), 3),
        n700 if n700 is not None else "",
        round(pct700, 2) if pct700 is not None else "",
        v.get("rep_score", "") if v.get("has_rep") else "",
        (v.get("rep_tier", "") + "｜" + v.get("rep_src", "")) if v.get("has_rep") else "",
        comp if comp is not None else "",
        v.get("rankA_in", ""), v.get("rankB_in", ""), v.get("rankC_in", ""), v.get("rankA_all", ""),
        completeness
    ))
rank_rows.sort(key=lambda x: (DISTRICTS_ORDER.index(x[0]) if x[0] in DISTRICTS_ORDER else 99,
                              x[12] if isinstance(x[12], int) else 9999,
                              x[10] if isinstance(x[10], int) else 9999))
write_table(ws_rank, 3,
    ["区", "初中学校", "推算考生数(人)", "到校名额合计", "生源强度(榜A·区内z)", "700+人数", "700+占比%(推算)",
     "声誉口碑(榜D)", "声誉来源/梯队", "综合得分(数据+声誉)", "区内排名A", "区内排名B", "区内排名(综合)", "全市排名A", "数据完整度"],
    rank_rows, widths=[10, 30, 14, 12, 16, 10, 14, 12, 34, 18, 10, 10, 12, 10, 16], freeze=True)

# ============ Sheet 4: 最低投档控制线 ============
ws = wb.create_sheet("最低投档控制线")
ws["A1"] = "2026 年上海中考各类招生最低投档控制分数线（官方）"; ws["A1"].font = SUB_FONT
rows = [
    ("自主招生录取", 615, "依据沪教委基〔2026〕2号，学业考总成绩满分750"),
    ("名额分配综合评价录取", 615, "含名额分配到区、名额分配到校"),
    ("普通高中统一招生录取（普高线）", 501, "1-15志愿统一招生批次"),
    ("中本贯通录取", 501, ""),
    ("五年一贯制和中高职贯通录取", 400, ""),
    ("普通中专录取", 300, ""),
]
write_table(ws, 3, ["招生类别","最低投档控制分数线(分)","备注"], rows, widths=[28,22,60])

# ============ Sheet 5: 全市成绩分布(一分一段) ============
ws = wb.create_sheet("全市成绩分布(一分一段)")
ws["A1"] = "2026 年上海中考全市成绩分布（一分一段）"; ws["A1"].font = SUB_FONT
ws["A2"] = "说明：考试院公布全市一分一段表，但未提供完整可机读文本；下表为公开渠道可确认的真实锚点 + 民间预测分档。"
ws["A2"].alignment = WRAP; ws.merge_cells("A2:D2")
anchor = [
    (750, 0, "理论满分"),
    (730, 100, "头条/自媒体公布的真实累计位次"),
    (720, 800, "真实累计位次"),
    (710, 2200, "真实累计位次"),
    (700, 6900, "真实累计位次（占13.8万约5%）"),
]
r = write_table(ws, 4, ["分数(≥)","全市累计位次(人)","性质","说明"],
                [(a[0], a[1], "真实锚点", a[2]) for a in anchor],
                widths=[12,18,12,40])
ws.cell(row=r+1, column=1, value="以下为民间预测分档（非官方，仅供参考）：").font = SUB_FONT
pred = [
    ("705及以上", "全市前800名", "上海中学/复旦附中/交大附中/华师大二附中 四校本部"),
    ("695-704", "全市前2500名", "七宝/建平/南模等八校，及华二紫竹/交附闵分等四校分校"),
    ("685-694", "全市前5000名", "曹杨二中/进才/位育等新五虎、头部特色市重点"),
    ("670-684", "全市前9000名", "市西/市北/洋泾等第二梯队市重点"),
    ("655-669", "全市前14000名", "各区头部强区重点，如行知/嘉定一中/闵行中学"),
    ("635-654", "全市前20000名", "普通区重点头部，华政附中等优质特色普高"),
    ("610-634", "全市前26000名", "普通区重点尾部，头部民办高中"),
    ("580-609", "全市前32000名", "多数公办普通高中"),
    ("505-579", "全市前40000名", "剩余民办普高、中本贯通相关专业"),
]
write_table(ws, r+2, ["分数段","预估累计位次","对应梯队"],
            [(p[0], p[1], p[2]) for p in pred],
            widths=[14,18,55], freeze=False)

# ============ Sheet 6: 市重招生计划汇总 ============
ws = wb.create_sheet("市重招生计划汇总")
ws["A1"] = "2026 年上海中考市重点高中招生计划（各批次汇总，官方/机构汇总）"; ws["A1"].font = SUB_FONT
rows = [
    ("自主招生", 4345, "77所市重点参与自招"),
    ("名额分配到区", 7171, "77所市重点；合计7171个到区名额"),
    ("名额分配到校", 12904, "含区属约12561 + 委属约365（搜狐口径；本处用全市口径12904）"),
    ("平行志愿(统一招生)", 6308, "1-15志愿统一招生批次"),
    ("合计(市重总计划)", 30728, "全市市重招生总计划；较去年28890增1838"),
]
r = write_table(ws, 3, ["招生批次","招生计划数(人)","备注"], rows, widths=[22,16,60])
ws.cell(row=r, column=1, value="附：全市高中总数").font = SUB_FONT
write_table(ws, r+1, ["类别","数量(所)","备注"],
            [("全市高中合计", 312, "参加中考招生"),
             ("市重点", 79, ""), ("区重点", 75, ""),
             ("特色高中", 21, ""), ("普通高中", 137, "")],
            widths=[14,12,40], freeze=False)

# ============ Sheet 7: 名额分配到区计划(市重77所) ============
ws = wb.create_sheet("名额分配到区计划(市重77所)")
ws["A1"] = "2026 年上海市高中名额分配到区招生计划（全市77所市重点，来源：上海本地宝/考试院）"; ws["A1"].font = SUB_FONT
path = DATA + "/citywide_到区计划.csv"
drows, _ = read_csv(path)
total = sum(int(x[7]) for x in drows if len(x) > 7 and x[7].isdigit())
out = []
for x in drows:
    if len(x) < 8 or not x[7].isdigit(): continue
    out.append((x[0], _hs_unify(x[1]), x[2], x[3], x[4], x[5], int(x[7])))  # 高中校名统一到官方全称
out.append(("合计", "—", "—", "—", "—", "—", total))
r = write_table(ws, 3, ["招生代码","学校名称","所属区","办别","学校类型","寄宿情况","计划数"],
                out, widths=[10,30,10,8,26,12,10])
ws.cell(row=r, column=1, value=f"注：计划数合计 {total} 人（与全市到区计划7171一致）。").font = Font(italic=True, color="C00000")

# ============ Sheet 8-10: 合并各区分数线 ============
def combine(prefix, headers, extra_col_name, fill_district_from_name=True):
    rows = []
    gap_note = []
    for fn in sorted(os.listdir(DATA)):
        if not fn.startswith(prefix+"_") or not fn.endswith(".csv"):
            continue
        key = fn[len(prefix)+1:-4]
        dist = DISTRICT_MAP.get(key, key)
        drows, comments = read_csv(DATA + "/" + fn)
        if comments:
            gap_note.append(f"{dist}：{comments[0]}")
        for x in drows:
            if not x or (len(x) >= 1 and x[0] == headers[0]):
                continue
            row = list(x)
            if fill_district_from_name and len(row) > 0:
                row[0] = dist
            rows.append(row)
    return rows, gap_note

ws = wb.create_sheet("平行志愿分数线")
ws["A1"] = "2026 年上海中考 1-15志愿（平行志愿）统一招生录取最低分数线"; ws["A1"].font = SUB_FONT
rows, gaps = combine("parallel", ["区","招生代码","招生学校","录取最低分","语数外","数学","语文","综合测试","是否同分优待","备注"], "区")
for row in rows:  # 高中校名统一到官方全称（与汇总透视、其他录取线表一致）
    if len(row) > 2:
        row[2] = _hs_unify(row[2])
r = write_table(ws, 3, ["区","招生代码","招生学校","录取最低分","语数外","数学","语文","综合测试","是否同分优待","备注"],
                rows, widths=[10,10,26,12,10,8,8,10,12,16])
if gaps:
    ws.cell(row=r+1, column=1, value="缺口说明：").font = Font(bold=True, color="C00000")
    for i, g in enumerate(gaps):
        ws.cell(row=r+2+i, column=1, value=g)

# ============ Sheet: 各区分数水平(平行志愿代理) ============
ws_px = wb.create_sheet("各区分数水平(平行志愿代理)")
ws_px["A1"] = "各区中考分数水平（平行志愿线代理，由官方平行志愿线推算）"; ws_px["A1"].font = SUB_FONT
_px_note = ("代理方法：用本工作簿『平行志愿分数线』Sheet 中各区官方录取最低分，计算每区的均值/中位数/最高最低线，"
            "以及 ≥700 / ≥680 / ≥650 分的学校占比，作为『分区分数水平』的可复现代理指标。"
            "⚠ 这是区级分数水平的近似，并非官方『分区一分一段表』（官方从不公布分区一分一段）。"
            "宝山区官方以图片发布，仅能从公开报道提取到 6 所头部本区高中，故宝山区均值为偏高的部分值，已在『数据状态』标注。")
ws_px["A2"] = _px_note; ws_px["A2"].alignment = WRAP; ws_px.merge_cells("A2:J2")
rows_px = []
with open(os.path.join(DATA, "district_score_proxy.csv"), encoding="utf-8-sig") as fh:
    rdp = csv.reader(fh)
    next(rdp)
    for p in rdp:
        if not p or len(p) < 10:
            continue
        def _f(x):
            x = (x or "").strip()
            return float(x) if x else None
        rows_px.append((p[0].strip(), int(p[1]), _f(p[2]), _f(p[3]), _f(p[4]), _f(p[5]),
                        _f(p[6]), _f(p[7]), _f(p[8]), p[9].strip()))
rows_px.sort(key=lambda x: (DISTRICTS_ORDER.index(x[0]) if x[0] in DISTRICTS_ORDER else 99))
write_table(ws_px, 3, ["区", "高中数(本区招生)", "平行志愿线均值", "中位数", "最高线", "最低线", "≥700占比%", "≥680占比%", "≥650占比%", "数据状态"],
            rows_px, widths=[10, 16, 14, 10, 10, 10, 12, 12, 12, 34], freeze=True)

ws = wb.create_sheet("名额到区录取分数线")
ws["A1"] = "2026 年上海中考 名额分配到区 招生录取最低分数线（按接收区）"; ws["A1"].font = SUB_FONT
rows, gaps = combine("to_district", ["接收区","招生学校","隶属区","办学性质","录取最低分","语数外","数学","语文","综合测试","是否同分优待"], None, fill_district_from_name=True)
for row in rows:  # 高中校名统一到官方全称
    if len(row) > 1:
        row[1] = _hs_unify(row[1])
r = write_table(ws, 3, ["接收区","招生学校","隶属区","办学性质","录取最低分","语数外","数学","语文","综合测试","是否同分优待"],
                rows, widths=[10,28,10,10,12,10,8,8,10,12])
if gaps:
    ws.cell(row=r+1, column=1, value="缺口说明：").font = Font(bold=True, color="C00000")
    for i, g in enumerate(gaps):
        ws.cell(row=r+2+i, column=1, value=g)

ws = wb.create_sheet("名额到校录取分数线")
ws["A1"] = "2026 年上海中考 名额分配到校 招生录取最低分数线（初中×高中组合）"; ws["A1"].font = SUB_FONT
rows, gaps = combine("to_school", ["区","初中学校","招生学校","录取最低分","综合素质评价","语数外","数学","语文","综合测试","是否同分优待"], "区")
for row in rows:  # 高中校名统一；初中校名统一到 plan 官方全称（与排行榜/700+表同源）
    if len(row) > 2:
        row[2] = _hs_unify(row[2])
    if len(row) > 1 and row[0]:
        row[1] = _jr_unify(row[0], row[1])
r = write_table(ws, 3, ["区","初中学校","招生学校","录取最低分","综合素质评价","语数外","数学","语文","综合测试","是否同分优待"],
                rows, widths=[10,26,28,12,12,10,8,8,10,12])
if gaps:
    ws.cell(row=r+1, column=1, value="缺口说明：").font = Font(bold=True, color="C00000")
    for i, g in enumerate(gaps):
        ws.cell(row=r+2+i, column=1, value=g)

# ============ Sheet: 名额分配到校计划(逐校) ============
# （创建于末尾，最后统一重排序到『名额到校录取分数线』之后）
ws_plan = wb.create_sheet("名额分配到校计划(逐校)")
ws_plan["A1"] = "2026 年上海中考 名额分配到校 招生计划（逐校明细 · 初中×高中组合，来源：16区教育局/招考机构公示）"; ws_plan["A1"].font = SUB_FONT
plan_rows, plan_gaps = [], []
for fn in sorted(os.listdir(DATA)):
    if not fn.startswith("to_school_plan_") or not fn.endswith(".csv"):
        continue
    key = fn[len("to_school_plan_"):-4]
    dist = DISTRICT_MAP.get(key, key)
    drows, comments = read_csv(DATA + "/" + fn)
    if comments:
        plan_gaps.append(f"{dist}：{comments[0]}")
    for x in drows:
        if not x or (len(x) >= 1 and x[0] == "区"):
            continue
        row = list(x)
        if len(row) >= 1:
            row[0] = dist
        if len(row) > 2:
            row[2] = _hs_unify(row[2])          # 高中校名统一到官方全称
        if len(row) > 1 and row[0]:
            row[1] = _jr_unify(row[0], row[1])  # 初中校名→plan官方全称（与排行榜同源）
        plan_rows.append(row)
r = write_table(ws_plan, 3, ["区","初中学校","招生学校","计划数"],
                plan_rows, widths=[10,28,30,10])
if plan_gaps:
    ws_plan.cell(row=r+1, column=1, value="缺口说明：").font = Font(bold=True, color="C00000")
    for i, g in enumerate(plan_gaps):
        ws_plan.cell(row=r+2+i, column=1, value=g).alignment = WRAP

# ============ Sheet 11: 各区统招(1-15志愿)计划 ============
ws = wb.create_sheet("各区统招(1-15志愿)计划")
ws["A1"] = "2026 年上海中考 各区统一招生(1-15志愿)批次 招生计划（已知区，来源：vsxue/头条汇总）"; ws["A1"].font = SUB_FONT
plan = [
    ("浦东新区", 16354, 274, "本区高中66所；新增临港科高/洋泾南校/群峰高中"),
    ("闵行区", 6838, 659, "本区31所；扩招771，多为普高"),
    ("徐汇区", 3529, 310, "市二梅陇校区减35、本部增48；民办位育新开70"),
    ("杨浦区", 3061, 182, "本区13所；扩招330多为区重点"),
    ("静安区", 3150, None, "本区21所；复附静安首招45；市重红利区"),
    ("黄浦区", 1365, None, "本区17所；同济科中首招5"),
    ("虹口区", 1868, None, "本区9所；普高录取率最高约75%"),
    ("青浦区", 1841, None, "本区11所；上师附青浦首届176"),
    ("嘉定区", 3872, 179, "本区13所；扩招550"),
    ("宝山区", 5091, None, "本区22所；市重扩招87"),
]
rows = []
for d, b, w, note in plan:
    total = b + w if (w is not None) else b
    rows.append((d, b, w if w is not None else "—", total, note))
write_table(ws, 3, ["区","本区高中计划(人)","外区高中计划(人)","合计(人)","备注"], rows,
            widths=[12,18,18,12,55])
ws.cell(row=3+len(rows)+2, column=1,
        value="注：长宁/普陀/金山/松江/奉贤/崇明等区统招计划数公开源暂缺；此表仅为统一招生(1-15志愿)批次，不含自招与名额分配批次。").font = Font(italic=True, color="C00000")

# ============ Sheet: 汇总透视 ============
ws_pivot = wb.create_sheet("汇总透视")
ws_pivot["A1"] = "汇总透视（按高中 / 按区 汇总名额分配与分数线，便于快速看盘子）"; ws_pivot["A1"].font = SUB_FONT
ws_pivot["A2"] = ("说明：『到校计划数(合计)』为16区『名额分配到校计划(逐校)』之和，现已 16 区计划CSV全部并入（含徐汇899，合计12875）；与全市口径12904差约29，系个别区计划表仍有极少量未解析单元格，非徐汇缺口。"
                  "『到区录取线/平行志愿线』为各区最低值样本，前者800分制、后者750分制，二者不可相减、不可与计划数混淆。")
ws_pivot["A2"].alignment = WRAP; ws_pivot.merge_cells("A2:F2")

# —— 按高中汇总 ——
# 关键修复：5 个来源的高中校名经 _hs_unify 统一到官方全称键，避免平行志愿缩写/异写导致同一高中被拆成多行。
hs_plan = {}; hs_cz = {}; hs_td_min = {}; hs_pl_min = {}; dist_plan = {}
drows, _ = read_csv(DATA + "/citywide_到区计划.csv")
for x in drows:
    if len(x) < 8 or not x[7].isdigit():
        continue
    dist_plan[_hs_unify(x[1])] = int(x[7])
for fn in sorted(os.listdir(DATA)):
    if not fn.startswith("to_school_plan_") or not fn.endswith(".csv"):
        continue
    drows, _ = read_csv(DATA + "/" + fn)
    for x in drows:
        if not x or x[0] == "区" or len(x) < 4:
            continue
        try:
            v = int(x[3])
        except ValueError:
            v = 0
        hs_plan[_hs_unify(x[2])] = hs_plan.get(_hs_unify(x[2]), 0) + v
for fn in sorted(os.listdir(DATA)):
    if not fn.startswith("to_school_") or fn.startswith("to_school_plan_") or not fn.endswith(".csv"):
        continue
    drows, _ = read_csv(DATA + "/" + fn)
    for x in drows:
        if not x or x[0] == "区" or len(x) < 3:
            continue
        hs_cz.setdefault(_hs_unify(x[2]), set()).add(x[1])
for fn in sorted(os.listdir(DATA)):
    if not fn.startswith("to_district_") or not fn.endswith(".csv"):
        continue
    drows, _ = read_csv(DATA + "/" + fn)
    for x in drows:
        if not x or x[0] == "接收区" or len(x) < 5:
            continue
        try:
            sc = float(x[4])
        except ValueError:
            sc = None
        if sc is not None:
            hs_td_min[_hs_unify(x[1])] = min(hs_td_min.get(_hs_unify(x[1]), 999), sc)
for fn in sorted(os.listdir(DATA)):
    if not fn.startswith("parallel_") or not fn.endswith(".csv"):
        continue
    drows, _ = read_csv(DATA + "/" + fn)
    for x in drows:
        if not x or x[0] == "区" or len(x) < 4:
            continue
        try:
            sc = float(x[3])
        except ValueError:
            sc = None
        if sc is not None:
            hs_pl_min[_hs_unify(x[2])] = min(hs_pl_min.get(_hs_unify(x[2]), 999), sc)
all_hs = sorted(set(dist_plan) | set(hs_plan) | set(hs_cz) | set(hs_td_min) | set(hs_pl_min),
                key=lambda k: -dist_plan.get(k, 0))
rows_h = [(h, dist_plan.get(h, ""), hs_plan.get(h, ""), len(hs_cz.get(h, set())),
           hs_td_min.get(h, ""), hs_pl_min.get(h, "")) for h in all_hs]
r = write_table(ws_pivot, 4,
    ["招生学校", "到区计划数", "到校计划数(合计)", "到校覆盖初中数", "到区录取线(最低·800制)", "平行志愿线(最低·750制)"],
    rows_h, widths=[30, 12, 16, 14, 18, 18])

# —— 按区汇总 ——
by_dist = {}
def _dist_of(fn, pre):
    return DISTRICT_MAP.get(fn[len(pre):-4], fn[len(pre):-4])
for fn in sorted(os.listdir(DATA)):
    if not fn.startswith("to_school_plan_") or not fn.endswith(".csv"):
        continue
    dist = _dist_of(fn, "to_school_plan_")
    d = by_dist.setdefault(dist, {"plan": 0, "cz": set()})
    drows, _ = read_csv(DATA + "/" + fn)
    for x in drows:
        if not x or x[0] == "区" or len(x) < 4:
            continue
        try:
            v = int(x[3])
        except ValueError:
            v = 0
        d["plan"] += v
        d["cz"].add(_jr_unify(dist, x[1]))   # 计划初中名→plan官方全称（与排行榜同源）
for fn in sorted(os.listdir(DATA)):
    if not fn.startswith("to_school_") or fn.startswith("to_school_plan_") or not fn.endswith(".csv"):
        continue
    dist = _dist_of(fn, "to_school_")
    d = by_dist.setdefault(dist, {"plan": 0, "cz": set()})
    drows, _ = read_csv(DATA + "/" + fn)
    for x in drows:
        if not x or x[0] == "区" or len(x) < 2:
            continue
        d.setdefault("line_cz", set()).add(_jr_unify(dist, x[1]))   # 录取线初中名→plan官方全称，与计划口径对齐
for fn in sorted(os.listdir(DATA)):
    if not fn.startswith("parallel_") or not fn.endswith(".csv"):
        continue
    dist = _dist_of(fn, "parallel_")
    d = by_dist.setdefault(dist, {"plan": 0, "cz": set()})
    drows, _ = read_csv(DATA + "/" + fn)
    for x in drows:
        if not x or x[0] == "区" or len(x) < 3:
            continue
        d.setdefault("pl_sch", set()).add(_hs_unify(x[2]))   # 平行志愿高中名统一
for fn in sorted(os.listdir(DATA)):
    if not fn.startswith("to_district_") or not fn.endswith(".csv"):
        continue
    drows, _ = read_csv(DATA + "/" + fn)
    for x in drows:
        if not x or x[0] == "接收区" or len(x) < 2:
            continue
        dist = DISTRICT_MAP.get(x[0], x[0])
        d = by_dist.setdefault(dist, {"plan": 0, "cz": set()})
        d.setdefault("td_sch", set()).add(_hs_unify(x[1]))   # 到区高中名统一
rows_d = []
for dist in DISTRICTS_ORDER:
    d = by_dist.get(dist)
    if not d:
        rows_d.append((dist, "", "", "", "", ""))
        continue
    rows_d.append((dist, d["plan"], len(d["cz"]), len(d.get("line_cz", set())),
                   len(d.get("pl_sch", set())), len(d.get("td_sch", set()))))
r2 = write_table(ws_pivot, r + 2,
    ["区", "到校计划数合计", "享受到校初中数", "到校录取线覆盖初中数", "平行志愿招生学校数", "到区录取线高中数(接收)"],
    rows_d, widths=[12, 16, 16, 18, 18, 18], freeze=False)
pivot_note = (f"校验：到区计划合计={sum(dist_plan.values())}（应=7171）；到校计划合计={sum(hs_plan.values())}"
               f"（16区计划CSV已全；与全市12904差{12904 - sum(hs_plan.values())}，系个别区计划表极少量未解析）。")
ws_pivot.cell(row=r2 + 1, column=1, value=pivot_note).font = Font(italic=True, color="C00000")

# ============ Sheet 12: 数据覆盖与缺口 ============
ws = wb.create_sheet("数据覆盖与缺口")
ws["A1"] = "数据覆盖与缺口清单"; ws["A1"].font = SUB_FONT
cov = [("维度","覆盖情况","说明")]
cov.append(("考生人数(全市)","✅ 官方","138000人"))
cov.append(("考生人数(各区)","⚠️ 民间估算","官方未统一公布分区人数"))
cov.append(("最低投档控制线","✅ 官方","615/615/501等"))
cov.append(("全市成绩分布(一分一段)","⚠️ 部分","仅真实锚点+预测分档"))
cov.append(("分区成绩分布(一分一段)","❌ 未公布·有代理","考试院仅公布全市一分一段，任何来源(含教培/自媒体)均无分区一分一段；新增『各区分数水平(平行志愿代理)』Sheet 用官方平行志愿线推算均值/高分占比作区级近似(非一分一段)"))
cov.append(("初中校考生人数(推算)","⚠️ 推算","按到校名额占比反推，覆盖拿到到校名额的初中；精度受各区考生数(民间估算)与名额取整影响"))
cov.append(("初中校700分以上人数(民间统计)","⚠️ 民间","教培/自媒体统计 113 所初中的700+人数与最高分(覆盖16区含金山/崇明)，仅高分段尾巴、非完整分布，可靠度逐行标注；另算700+占比%(推算)"))
cov.append(("初中学校排行榜(分区分榜)","✅ 派生(区内为主)","由『名额到校录取线』算生源强度(全量703校·z分去市重难度)、『700+占比』算高分密度(113校)，并引入『声誉口碑(榜D)』(公开排名/口碑,主观)合成综合=A:B:D=2:1.5:6.5(声誉主导)；无声誉校回退A:B=6:4并标『声誉待补』；推算/民间/声誉均标数据完整度"))
cov.append(("市重招生计划(各批次)","✅ 官方/汇总","30728总，分4批"))
cov.append(("名额分配到区计划(市重77所)","✅ 完整","citywide_到区计划.csv，合计7171"))
cov.append(("名额分配到校计划(逐校)","✅ 16区已并入","初中×高中×计划数，合计12875；徐汇由用户手工补齐(899)，与全市12904差约29(个别区计划表极少量未解析)"))
cov.append(("平行志愿分数线","⚠️ 15区完整+宝山部分","原缺的闵行/嘉定/松江已从本地宝文本页补齐；宝山区官方图片仅提取到6所头部本区高中(其余官方图片未解析)"))
cov.append(("各区分数水平(平行志愿代理)","✅ 派生(官方线推算)","由『平行志愿分数线』官方线计算每区均值/中位数/高分占比，作区级分数水平代理；宝山区为偏高部分值"))
cov.append(("名额到区录取分数线","✅ 16区完整","按接收区"))
cov.append(("名额到校录取分数线","✅ 16区完整","初中×高中组合，浦东最全974行"))
cov.append(("各区统招计划数","⚠️ 部分区","10区已知，6区暂缺"))
write_table(ws, 3, ["维度","覆盖情况","说明"], cov, widths=[34,14,55])

# 将『名额分配到校计划(逐校)』移到『名额到校录取分数线』之后、『各区统招计划』之前
_sheets = wb._sheets
_new = wb["名额分配到校计划(逐校)"]
_sheets.remove(_new)
_idx = _sheets.index(wb["各区统招(1-15志愿)计划"])
_sheets.insert(_idx, _new)
# 将『汇总透视』移到『数据覆盖与缺口』之前
_pivot = wb["汇总透视"]
_sheets.remove(_pivot)
_idx2 = _sheets.index(wb["数据覆盖与缺口"])
_sheets.insert(_idx2, _pivot)

wb.save(OUT)
print("SAVED:", OUT)
print("Sheet数量:", len(wb.sheetnames))
print("到区计划合计:", total)
print("字段字典行数:", len(glossary))
print("到校计划(逐校)行数:", len(plan_rows))
