#!/usr/bin/env python3
"""
ETL脚本：提取4个缺失区的2025年名额分配到校数据

处理区县:
1. 普陀区 - 横向表格，两行表头
2. 金山区 - 横向表格，简单格式
3. 虹口区 - XLSX格式，两行表头

注意: 徐汇区PDF包含的是录取分数线数据，不是名额分配数据

输出CSV格式：
year,district,district_code,middle_school_name,high_school_code,high_school_name,quota_count,source
"""

import csv
import os
import re
from pathlib import Path
from datetime import datetime

try:
    import pdfplumber
except ImportError:
    print("请先安装pdfplumber: pip install pdfplumber")
    exit(1)

try:
    import openpyxl
except ImportError:
    print("请先安装openpyxl: pip install openpyxl")
    exit(1)

# ============================================================================
# 配置
# ============================================================================
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR = os.path.join(BASE_DIR, "raw", "2025", "quota_district")
PROCESSED_DIR = os.path.join(BASE_DIR, "processed", "2025", "quota_school")

Path(PROCESSED_DIR).mkdir(parents=True, exist_ok=True)

# 已知高中代码到名称映射
KNOWN_HIGH_SCHOOLS = {
    '042032': '上海市上海中学',
    '102056': '上海交通大学附属中学',
    '102057': '复旦大学附属中学',
    '152003': '华东师范大学第二附属中学',
    '152006': '上海师范大学附属中学',
    # 普陀区高中
    '072001': '上海市晋元高级中学',
    '072002': '上海市曹杨第二中学',
    '073003': '上海市宜川中学',
    '073082': '华东师范大学第二附属中学（普陀校区）',
    # 金山区高中
    '162000': '上海市金山中学',
    '163002': '华东师范大学第三附属中学',
    # 虹口区高中
    '092001': '复旦大学附属复兴中学',
    '092002': '华东师范大学第一附属中学',
    '093001': '上海财经大学附属北郊高级中学',
}

# 各区高中代码列表（按表格顺序）
DISTRICT_HIGH_SCHOOLS = {
    '普陀区': ['042032', '102057', '102056', '152003', '152006', '073082', '072002', '072001', '073003'],
    '金山区': ['042032', '102056', '102057', '152003', '152006', '162000', '163002'],
    '虹口区': ['092001', '092002', '093001', '042032', '152003', '102057', '102056'],
}


def clean_text(text):
    """清理文本，去除换行和多余空格"""
    if not text:
        return ""
    return str(text).replace('\n', '').replace('\r', '').strip()


def extract_code_from_text(text):
    """从文本中提取6位数字代码"""
    if not text:
        return None
    text = clean_text(text)
    match = re.search(r'(\d{6})', text)
    if match:
        return match.group(1)
    return None


def is_quota_number(value):
    """检查是否是有效的名额数字"""
    if value is None:
        return False
    try:
        num = int(str(value).strip())
        return 0 <= num <= 500
    except (ValueError, TypeError):
        return False


def extract_putuo_format(pdf_path, district, district_code):
    """
    普陀区格式：横向表格（两行表头）
    行0: 学校代码 | 学校名称 | 委属市示范性高中分配结果 | ... | 区属市示范性高中分配结果
    行1: | | 华二 | 上中 | 复附 | 交附 | 上师大 | 华二普陀 | 曹杨二中 | 晋元 | 宜川
    行2+: 071045 | 上海市洵阳中学 | 0 | 0 | 0 | 0 | 0 | 1 | 2 | 1 | 1
    """
    records = []
    high_codes = DISTRICT_HIGH_SCHOOLS.get(district, [])

    print(f"  解析普陀区格式")

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 3:
                    continue

                # 从第三行开始解析数据
                for row in table[2:]:
                    if not row or len(row) < 4:
                        continue

                    # 提取初中名称（清理乱码前缀）
                    middle_name = clean_text(row[1]) if len(row) > 1 else ""
                    middle_name = re.sub(r'^(未|经|许|得|，|允|不|转|载|可|院)\s*', '', middle_name)

                    if not middle_name or '中学' not in middle_name:
                        continue
                    if '合计' in middle_name:
                        continue

                    # 提取名额（从第3列开始，对应高中代码）
                    for col_idx, high_code in enumerate(high_codes):
                        data_col = col_idx + 2  # 数据从第3列开始
                        if data_col < len(row) and is_quota_number(row[data_col]):
                            quota = int(str(row[data_col]).strip())
                            if quota > 0:
                                high_name = KNOWN_HIGH_SCHOOLS.get(high_code, f"高中{high_code}")
                                records.append({
                                    'year': 2025, 'district': district, 'district_code': district_code,
                                    'middle_school_name': middle_name, 'high_school_code': high_code,
                                    'high_school_name': high_name, 'quota_count': quota, 'source': f'page_{page_num}'
                                })

    print(f"    提取 {len(records)} 条记录")
    return records


def extract_jinshan_format(pdf_path, district, district_code):
    """
    金山区格式：横向表格（简单格式）
    行0: 序号 | 学校简称 | 委属... | 金山中学... | 华师大三附中...
    行1+: 1 | 上海市罗星中学 | | 19 | 20
    """
    records = []
    high_codes = DISTRICT_HIGH_SCHOOLS.get(district, [])

    print(f"  解析金山区格式")

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue

                # 从第二行开始解析数据
                for row in table[1:]:
                    if not row or len(row) < 3:
                        continue

                    # 提取初中名称
                    middle_name = clean_text(row[1]) if len(row) > 1 else ""

                    if not middle_name or '中学' not in middle_name:
                        continue
                    if '合计' in middle_name or '学校' in middle_name:
                        continue

                    # 提取名额（从第3列开始）
                    for col_idx, high_code in enumerate(high_codes):
                        data_col = col_idx + 2  # 数据从第3列开始
                        if data_col < len(row) and is_quota_number(row[data_col]):
                            quota = int(str(row[data_col]).strip())
                            if quota > 0:
                                high_name = KNOWN_HIGH_SCHOOLS.get(high_code, f"高中{high_code}")
                                records.append({
                                    'year': 2025, 'district': district, 'district_code': district_code,
                                    'middle_school_name': middle_name, 'high_school_code': high_code,
                                    'high_school_name': high_name, 'quota_count': quota, 'source': f'page_{page_num}'
                                })

    print(f"    提取 {len(records)} 条记录")
    return records


def extract_hongkou_xlsx(xlsx_path, district, district_code):
    """
    虹口区格式：XLSX文件（两行表头）
    行1: 序号 | 初中学校 | 复旦 | 华师大 | 上财 | 委属"四校" | | |
    行2: | | 复兴 | 一附中 | 北郊 | 上中 | 华二 | 复附 | 交附
    行3+: 1 | 虹口实验 | 14 | 13 | 10 | | | |
    """
    records = []
    high_codes = DISTRICT_HIGH_SCHOOLS.get(district, [])

    print(f"  解析虹口区XLSX格式")

    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb.active

    # 从第三行开始解析数据
    for row_idx in range(3, sheet.max_row + 1):
        # 提取初中名称
        middle_name = sheet.cell(row=row_idx, column=2).value
        if not middle_name:
            continue
        middle_name = clean_text(middle_name)

        if not middle_name or '中学' not in middle_name:
            # 可能是简称，尝试添加"中学"
            if middle_name and len(middle_name) >= 2:
                middle_name = middle_name + '中学' if not middle_name.endswith('中学') else middle_name
            else:
                continue

        if '合计' in middle_name:
            continue

        # 提取名额（从第3列开始）
        for col_idx, high_code in enumerate(high_codes):
            data_col = col_idx + 3  # 数据从第3列开始
            cell_value = sheet.cell(row=row_idx, column=data_col).value
            if is_quota_number(cell_value):
                quota = int(str(cell_value).strip())
                if quota > 0:
                    high_name = KNOWN_HIGH_SCHOOLS.get(high_code, f"高中{high_code}")
                    records.append({
                        'year': 2025, 'district': district, 'district_code': district_code,
                        'middle_school_name': middle_name, 'high_school_code': high_code,
                        'high_school_name': high_name, 'quota_count': quota, 'source': 'xlsx'
                    })

    print(f"    提取 {len(records)} 条记录")
    return records


def save_to_csv(records, district, district_code):
    """保存数据到CSV文件"""
    if not records:
        print(f"  警告：{district}没有提取到任何记录")
        return None

    output_file = os.path.join(PROCESSED_DIR, f"{district}_2025_名额分配到校.csv")

    with open(output_file, 'w', encoding='utf-8', newline='') as f:
        fieldnames = ['year', 'district', 'district_code', 'middle_school_name',
                      'high_school_code', 'high_school_name', 'quota_count', 'source']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)

    print(f"  输出文件: {output_file}")
    return output_file


def main():
    print("=" * 70)
    print("2025年名额分配到校数据提取 - 4个缺失区")
    print(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    results = []

    # 1. 处理普陀区
    print("\n处理普陀区:")
    putuo_path = os.path.join(RAW_DIR, "2025年名额分配到校-普陀.pdf")
    if os.path.exists(putuo_path):
        records = extract_putuo_format(putuo_path, '普陀区', 'PT')
        output_file = save_to_csv(records, '普陀区', 'PT')
        results.append({'district': '普陀区', 'record_count': len(records)})
    else:
        print(f"  文件不存在: {putuo_path}")

    # 2. 处理金山区
    print("\n处理金山区:")
    jinshan_path = os.path.join(RAW_DIR, "2025年名额分配到校-金山.pdf")
    if os.path.exists(jinshan_path):
        records = extract_jinshan_format(jinshan_path, '金山区', 'JS')
        output_file = save_to_csv(records, '金山区', 'JS')
        results.append({'district': '金山区', 'record_count': len(records)})
    else:
        print(f"  文件不存在: {jinshan_path}")

    # 3. 处理虹口区 (XLSX)
    print("\n处理虹口区:")
    hongkou_path = os.path.join(RAW_DIR, "2025年名额分配到校-虹口.xlsx")
    if os.path.exists(hongkou_path):
        records = extract_hongkou_xlsx(hongkou_path, '虹口区', 'HK')
        output_file = save_to_csv(records, '虹口区', 'HK')
        results.append({'district': '虹口区', 'record_count': len(records)})
    else:
        print(f"  文件不存在: {hongkou_path}")

    # 4. 徐汇区说明
    print("\n处理徐汇区:")
    print("  跳过: 徐汇区PDF包含的是录取分数线数据，不是名额分配数据")
    print("  需要查找徐汇区名额分配到校的原始PDF文件")

    # 汇总结果
    print("\n" + "=" * 70)
    print("处理结果汇总")
    print("=" * 70)

    total_records = 0
    for result in results:
        print(f"{result['district']}: {result['record_count']} 条记录")
        total_records += result['record_count']

    print(f"\n总记录数: {total_records}")
    print(f"输出目录: {PROCESSED_DIR}")


if __name__ == '__main__':
    main()
