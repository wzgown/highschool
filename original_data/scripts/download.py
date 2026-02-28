import os
import requests
from openpyxl import load_workbook

# 配置参数
EXCEL_FILE = "2024年上海市高中学校自主招生录取方案.xlsx"
SHEET_NAME = "Table 1"
PDF_DIR = "自主招生录取方案"

# 创建保存目录
os.makedirs(PDF_DIR, exist_ok=True)

# 加载Excel文件
wb = load_workbook(EXCEL_FILE)
ws = wb[SHEET_NAME]

# 遍历数据行（从第2行开始）
for row in ws.iter_rows(min_row=2):
    # 提取关键字段
    code = str(row[1].value).strip()  # B列：招生代码
    name = str(row[4].value).strip()  # E列：学校名称
    link_cell = row[6]                # G列：超链接单元格

    # 获取PDF链接
    if not link_cell.hyperlink:
        print(f"⚠️ 跳过第 {row[0].row} 行：未找到下载链接")
        continue
    pdf_url = link_cell.hyperlink.target

    # 生成文件名（自动补全6位招生代码）
    formatted_code = code.zfill(6)  # 将代码补零至6位
    filename = f"{formatted_code}-{name}.pdf"
    save_path = os.path.join(PDF_DIR, filename)

    # 下载文件
    try:
        res = requests.get(pdf_url, timeout=10)
        res.raise_for_status()
        with open(save_path, "wb") as f:
            f.write(res.content)
        print(f"✅ 已下载：{filename}")
    except Exception as e:
        print(f"❌ 下载失败：{filename} | 错误：{str(e)}")

wb.close()
print("\n全部处理完成，请检查输出目录。")
